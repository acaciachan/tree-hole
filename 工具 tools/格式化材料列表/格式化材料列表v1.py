import csv
import glob
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ===== 用户可调节参数 =====
input_lang = "zh_cn"  # 原始材料列表语言
output_lang = ["zh_cn"]  # 排版后的材料列表的语言，可以填多个："minecraft", "en_us", "zh_cn"
sort_template = "list.csv"  # 排序模板文件名
units_def = ["箱", 54, "盒", 27, "组", 64, "个"]  # 单位定义，数字为两两之间倍数，可随意增添删减修改
input_dir = "."  # 输入目录
output_dir = "."  # 输出目录
combine_limit = 12  # 空行合并设置：False=不合并，True=合并所有，数字=智能合并（最小值为1）
target_columns = 4  # 分列数量，默认为4列

# ===== 行类型判断函数 =====
def is_title(row):
    """判断一行是否为标题行（空行或以@开头的行）
    参数:
        row: 要检查的行数据（列表）
    返回:
        bool: 如果是标题行则返回True，否则返回False
    """
    if not row:  # 空列表直接返回True
        return True
    first_cell = row[0] if len(row) > 0 else None
    return (
        first_cell is None  # None视为标题行
        or (isinstance(first_cell, str) and (
            first_cell.strip() == ""  # 空字符串视为标题行
            or first_cell.startswith('@')  # 以@开头的字符串视为标题行
        ))
    )

def is_separator_title(row):
    """判断一行是否为分隔标题行（任意元素包含@符号）
    参数:
        row: 要检查的行数据（列表）
    返回:
        bool: 如果是分隔标题行则返回True，否则返回False
    """
    if not row:
        return False
    return any(
        isinstance(cell, str) and cell.startswith('@')
        for cell in row
    )

# ===== 单位处理函数 =====
def get_units(units_list):
    """处理单位定义列表，计算每个单位的实际倍数
    参数:
        units_list: 单位定义列表，格式为[单位名, 倍数, 单位名, 倍数,...]
    返回:
        list: 处理后的单位列表，包含单位名和对应的总倍数
    """
    units_pro = []
    index = 0
    for item in units_list: 
        if isinstance(item, str):
            units_pro.append(item)
        elif isinstance(item, int):
            num = 1
            for item2 in units_list[index::2]:
                num *= item2
            units_pro.append(num)
        else: 
            return None
        index += 1
    units_pro.append(1)    # 添加基础单位倍数1
    return units_pro

# 初始化单位处理结果
units_pro = get_units(units_def)
print("单位列表（单位后面的数字代表此单位和最小单位的倍数）：")
print(units_pro)

# ===== 数量格式化函数 =====
def formatted_num(num):
    """将数字格式化为带单位的字符串（如：2箱3盒5个）
    参数:
        num: 要格式化的数字
    返回:
        str: 格式化后的字符串，如"2箱3盒5个"
    """
    if num is None:
        return None
    elif num == "":
        return ""
    elif num == 0:
        return "0" + units_pro[-2]  # 返回0加最小单位
    
    formatted_num = ""
    skip_zero = True # 标记是否跳过开头的零数量单位

    for i in range(len(units_pro)//2):
        quotient = num // units_pro[i*2+1] # 计算当前单位的数量
        remainder = num % units_pro[i*2+1] # 计算余数
        
        if quotient == 0 and skip_zero:
            num = remainder
            continue # 跳过前面的零数量单位
        
        # 计算需要补零的数量（保持数字位数一致）
        if i == 0:
            zero_count = 0 # 第一个单位不需要补充零
        else:
            zero_count = max(0, len(str(units_def[i*2-1])) - len(str(quotient)))
        
        if skip_zero == True: # 处理第一个非零单位
            zero_count = 0 # 不补充零
            skip_zero = False # 后续单位不再跳过

        formatted_num += '0' * zero_count + str(quotient) + units_pro[i*2] # 拼接单位
        num = remainder # 更新余数
    
    return formatted_num

# ===== 导入排序模板CSV文件为列表 =====
def get_csv_as_a_list(input_file, selected_columns):
    """从排序模板CSV文件中读取数据（自动处理BOM）
    参数:
        input_file: 输入CSV文件路径
        selected_columns: 需要选择的列名列表
    返回:
        list: 包含CSV数据的二维列表
    """
    result = []
    try:
        with open(input_file, mode='r', encoding='utf-8-sig') as infile:  # 关键修改：utf-8-sig
            reader = csv.DictReader(infile)
            # 检查语言列是否存在
            missing_columns = [col for col in selected_columns if col not in reader.fieldnames]
            if missing_columns:
                raise ValueError(f"排序模板缺少语言列: {missing_columns}")
            
            # 添加表头和数据行
            result.append(selected_columns)
            for row in reader:
                result.append([row.get(col, "") for col in selected_columns])
    except Exception as e:
        print(f"读取排序模板错误: {str(e)}")
        raise
    return result

# ===== 导入材料列表CSV文件为列表 =====
def get_material_list(file_path):
    """从材料CSV文件中读取材料列表（自动处理BOM）
    参数:
        file_path: 材料CSV文件路径
    返回:
        list: 包含材料名和数量的列表，格式为[[材料名, 数量], ...]
    """
    result = []
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as file:  # 关键修改：utf-8-sig
            csv_reader = csv.reader(file)
            next(csv_reader)  # 跳过标题行
            for row in csv_reader:
                if len(row) >= 3:
                    # 可选：移除潜在的BOM残留（极端情况）
                    material_name = row[0].lstrip('\ufeff') if row[0] else row[0]
                    result.append([material_name, int(row[2])])
    except FileNotFoundError:
        print(f"错误: 文件 {file_path} 未找到")
    except Exception as e:
        print(f"读取材料文件错误: {str(e)}")
    return result   

# ===== 过滤数据列表 =====
def filter_data_list(data):
    """过滤数据列表，移除空行和无效数据
    参数:
        data: 原始数据列表
    返回:
        表头, 过滤后的数据
    """
    output_list_head = data[0]
    output_list = []
    last_was_empty = False  # 标记上一行是否为标题行
    
    for row in data[1:]:
        if is_title(row):
            if not last_was_empty or is_separator_title(row):  # 保留作为块标题的标题行
                output_list.append(row)
            last_was_empty = True
        else:
            if row[-1] is not None:  # 只保留有数量的行
                output_list.append(row)
                last_was_empty = False

    return output_list_head, output_list

# ===== 智能合并空行分隔 =====
def combine_blocks(output_list, combine_limit, num_columns):
    """根据设置合并小块数据（不合并含有@符号的分隔标题行）
    参数:
        output_list: 要处理的数据列表
        combine_limit: 合并限制参数
        num_columns: 列数
    返回:
        list: 合并后的数据列表
    """
    if combine_limit is False:  # 不合并
        return output_list
    
    if isinstance(combine_limit, (int, float)):
        if combine_limit < 1:
            print(f"警告: combine_limit值 {combine_limit} 无效，已自动设置为1")
            combine_limit = 1
    
    # 获取所有分隔行的索引
    block_indices = [i for i, row in enumerate(output_list) if is_title(row)]
    if not block_indices:  # 没有可合并的空行分隔
        return output_list

    # 根据空行分隔将数据分成多个块
    blocks = []
    for i, block_start in enumerate(block_indices):
        block_end = block_indices[i+1] if i+1 < len(block_indices) else len(output_list)
        blocks.append(output_list[block_start:block_end])

    # 如果设置为合并所有
    if combine_limit is True:
        combined = []
        for block in blocks:
            combined.append(block)
        return combined
    
    # 智能合并小块
    i = 0
    while i < len(blocks) - 1:
        current_block = blocks[i]
        next_block = blocks[i + 1]

        # 跳过含有@符号的分隔行
        if (is_separator_title(next_block[0])):
            i += 1
            continue
        
        current_size = len(current_block) - 1  # 当前块的项目数
        next_size = len(next_block) - 1        # 下一块的项目数
        
        # 如果两个块都小于等于限制且合并后不超过两倍限制
        if (current_size <= combine_limit):
            
            merged_block = current_block + next_block[1:]  # 合并块
            blocks[i] = merged_block
            del blocks[i + 1]
        else:
            i += 1
    
    # 将合并后的块重新组合成列表
    combined_list = []
    for block in blocks:
        combined_list.append(block)

    return combined_list

def split_into_columns(output_list, output_list_head, target_cols):
    """将数据列表分割成多列（每个块作为一个整体分配到列）
    参数:
        output_list: 要分割的数据列表（每个元素是一个块）
        output_list_head: 表头列表
        target_cols: 目标列数
    返回:
        tuple: (表头列表, 分割后的列数据)
    """
    if target_cols < 1:
        print(f"警告: target_columns值 {target_cols} 无效，已自动设置为1")
        target_cols = 1
    
    if target_cols == 1:  # 1列相当于不需要分列
        return output_list_head, [output_list]

    # 计算每列应该包含的块数
    blocks_per_col = len(output_list) // target_cols
    remainder = len(output_list) % target_cols
    
    # 分配块到各列
    cols = []
    start = 0
    for _ in range(target_cols):
        end = start + blocks_per_col + (1 if remainder > 0 else 0)
        cols.append(output_list[start:end])
        start = end
        remainder -= 1

    return output_list_head, cols

def process_chunk(output_list):
    """在所有含有@符号的分隔标题行下面插入一个空行
    参数:
        output_list: 要处理的数据列表（每个元素是一个块）
    返回:
        list: 处理后的数据列表
    """
    new_list = []
    
    for block in output_list:
        new_block = []
        
        if len(block) == 1: # 如果块只有一行标题，就不添加此行
            continue
        
        for row in block:
            new_block.append(row)
            if is_separator_title(row): # 如果是分隔标题行，就在下面添加一个空行
                new_block.append([None] * len(row))  # 插入空行
        
        new_list.append(new_block)
    
    return new_list

def export_to_excel(output_list_head, columns, filename):
    """将处理后的数据导出到Excel文件
    参数:
        output_list_head: 表头列表，决定每个大列有多少列
        columns: 列数据（每个元素是一个大列，包含多个块）
        filename: 输出文件名
    """
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.freeze_panes = 'A2'  # 冻结首行
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    block_title_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 浅灰色
    separate_title_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # 灰色
    right_alignment = Alignment(horizontal="right", vertical="center")

    # 写入表头
    col_offset = 0
    for _ in range(len(columns)):
        for i, col_name in enumerate(output_list_head):
            cell = ws.cell(row=1, column=col_offset + i + 1, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        col_offset += len(output_list_head) + 1  # 列间留空

    # 写入数据
    col_offset = 0
    for col_blocks in columns:  # 遍历每个大列
        # 首先收集这个大列中所有块的所有行
        all_rows = []
        for block in col_blocks:  # 遍历每个块
            all_rows.extend(block)  # 将块中的所有行加入
        
        # 写入这个大列的所有行
        for row_num, row_data in enumerate(all_rows):
            # 检查是否是标题行
            is_title_flag = is_title(row_data)
            is_separate_flag = is_separator_title(row_data)
            
            # 写入该行的每个单元格
            for col_num in range(len(output_list_head)):  # 根据表头决定列数
                value = row_data[col_num] if col_num < len(row_data) else None
                cell = ws.cell(row=row_num + 2, column=col_offset + col_num + 1, value=value)
                
                # 应用样式
                if is_separate_flag:
                    cell.fill = separate_title_fill
                elif is_title_flag:
                    cell.fill = block_title_fill
                
                # 数量列右对齐
                col_name = output_list_head[col_num]
                if col_name in ["数量", "格式化数量"]:
                    cell.alignment = right_alignment
        
        # 调整当前大列的列宽
        for col_num in range(len(output_list_head)):
            relative_col = col_offset + col_num + 1
            col_name = output_list_head[col_num]
            if col_name == "数量":
                ws.column_dimensions[get_column_letter(relative_col)].width = 9
            elif col_name == "格式化数量":
                ws.column_dimensions[get_column_letter(relative_col)].width = 18
            else:
                ws.column_dimensions[get_column_letter(relative_col)].width = 18
        
        col_offset += len(output_list_head) + 1  # 列间留空

    # 保存文件
    output_path = os.path.join(output_dir, filename.replace('.csv', '.xlsx'))
    wb.save(output_path)
    print(f"文件已成功导出到：{output_path}")

# ===== 主处理函数 =====
def process_material_file(material_file):
    """处理单个材料文件的主函数
    参数:
        material_file: 材料文件名
    """
    print(f"\n处理文件: {material_file}")
    
    # 确定需要选择的列
    extra_lang = input_lang not in output_lang
    selected_columns = output_lang + [input_lang] if extra_lang else output_lang.copy()
    
    # 读取排序模板
    sort_path = os.path.join(input_dir, sort_template)
    data_list = get_csv_as_a_list(sort_path, selected_columns)
    
    # 读取材料列表
    material_path = os.path.join(input_dir, material_file)
    original_list = get_material_list(material_path)
    
    # 获取输入语言的列索引
    target_col_index = data_list[0].index(input_lang)
    
    # 添加数量列
    data_list[0].append("数量")
    for i in range(1, len(data_list)):
        data_list[i].append(None)
    
    # 创建材料名到数量的映射
    material_quantity_map = {item[0]: item[1] for item in original_list}
    processed_materials = set()  # 记录已处理的材料
    last_col_index = len(data_list[0]) - 1  # 数量列的索引

    # 填充数量数据
    for row in data_list[1:]:
        if not row:
            continue
        
        material_name = row[target_col_index]
        if material_name not in processed_materials and material_name in material_quantity_map:
            row[last_col_index] = material_quantity_map[material_name]
            processed_materials.add(material_name)

    # 收集缺失的材料（存在于原始列表但不在模板中）
    missing_materials = []
    for material, quantity in original_list:
        if material not in processed_materials:
            # 创建新行：output_lang列填充为input_lang的名称，数量列填充实际数量
            new_row = [material]*(len(selected_columns))
            new_row.append(quantity)
            missing_materials.append(new_row)

    # 将缺失材料插入到data_list中（假设最后一行已经是"@排序模板缺失的材料"）
    if missing_materials:
        # 找到最后一个标题行的位置（包含@符号的行）
        last_title_index = len(data_list) - 1
        while last_title_index >= 0 and not is_separator_title(data_list[last_title_index]):
            last_title_index -= 1
        
        if last_title_index >= 0:  # 确保找到了标题行
            # 在标题行后面插入缺失材料
            data_list[last_title_index+1:last_title_index+1] = missing_materials

    # 如果需要移除输入语言列
    if extra_lang:
        lang_index = data_list[0].index(input_lang)
        del data_list[0][lang_index]
        for row in data_list[1:]:
            del row[lang_index]

    # 过滤数据
    output_list_head, output_list = filter_data_list(data_list)

    # 添加格式化数量列
    output_list_head.append("格式化数量")
    for row in output_list:
        row.append(formatted_num(row[-1]))

    # 合并小块
    output_list = combine_blocks(output_list, combine_limit, len(output_list_head))

    # 在所有分隔行下面插入空行，移除无内容的分隔行
    output_list = process_chunk(output_list)

    # 分列
    output_list_head, output_list = split_into_columns(output_list, output_list_head, target_columns)

    # 导出到Excel
    export_to_excel(output_list_head, output_list, material_file)

# ===== 主程序 =====
if __name__ == "__main__":
    """程序主入口，查找并处理所有匹配的材料文件"""
    print(f"当前工作目录: {os.getcwd()}")
    print(f"搜索路径: {os.path.join(input_dir, 'material_*.csv')}")
    
    # 查找所有以'material_'开头的CSV文件
    material_files = glob.glob(os.path.join(input_dir, "material_*.csv"))
    
    print(f"找到的文件: {material_files}")
    
    if not material_files:
        print(f"在目录 '{input_dir}' 中没有找到以'material_'开头的CSV文件")
        # 列出目录内容帮助调试
        print("目录内容:", os.listdir(input_dir))
    else:
        for material_file in material_files:
            file_name = os.path.basename(material_file)
            process_material_file(file_name)
    
    print("\n所有文件处理完成！")